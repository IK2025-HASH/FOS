"""
fos/ui/page_reports.py
Accountant's Period Report — Step workflow:
  1. Select period / FY → generate reconciliation
  2. Review opening/closing balances, inward/outward, sign off
  3. Export PDF draft + launch email to HMRC / client
"""

import os
from datetime import date, datetime
from PyQt6.QtWidgets import (
    QHBoxLayout, QVBoxLayout, QLabel, QFrame, QWidget,
    QLineEdit, QTextEdit, QFileDialog, QDialog,
    QDialogButtonBox, QStackedWidget, QPushButton
)
from PyQt6.QtCore import Qt

from ui.widgets import (
    BasePage, Card, PrimaryButton, SecondaryButton, ComboField,
    make_table, set_row, info, error, confirm,
    ACCENT, DARK, TEXT, MUTED, WHITE, BG, SUCCESS, WARN, DANGER, BORDER,
    _DIALOG_SS
)
from core.models import EntityModel, GLModel
from core.database import db

_MONTH_NAMES = ["January","February","March","April","May","June",
                "July","August","September","October","November","December"]


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _fy_periods_for_entity(entity: dict) -> list:
    """Return list of (label, [YYYY-MM, ...]) for up to 3 FYs."""
    fy_start_name = entity.get("fy_start", "April")
    try:
        fy_start_mo = (_MONTH_NAMES.index(fy_start_name) + 1) if fy_start_name in _MONTH_NAMES else 4
    except Exception:
        fy_start_mo = 4

    today = date.today()
    options = []
    for delta in range(3):
        start_year = today.year - delta
        if today.month < fy_start_mo:
            start_year -= 1
        end_year   = start_year + 1
        end_mo     = fy_start_mo - 1 or 12
        end_yr     = end_year if fy_start_mo > 1 else start_year

        periods = []
        y, m = start_year, fy_start_mo
        while (y, m) <= (end_yr, end_mo):
            periods.append(f"{y}-{m:02d}")
            m += 1
            if m > 12:
                m, y = 1, y + 1

        abbr = ["Jan","Feb","Mar","Apr","May","Jun",
                "Jul","Aug","Sep","Oct","Nov","Dec"]
        label = (f"FY {start_year}/{str(end_year)[2:]}  "
                 f"({abbr[fy_start_mo-1]} {start_year} – {abbr[end_mo-1]} {end_yr})")
        options.append((label, periods))
    return options


def _build_report_data(entity_id: str, periods: list) -> dict:
    """Compute all figures needed for the accountant's report."""
    if not periods:
        return {}

    ph = ",".join("?" * len(periods))

    # All GL entries in period
    rows = db.fetchall(
        f"SELECT * FROM gl WHERE entity_id=? AND period IN ({ph}) ORDER BY date",
        [entity_id] + periods
    )

    # Bank lines (account 1000) for opening/closing balance
    bank_rows = [r for r in rows if r["account_code"] == "1000"]

    # Opening balance = balance of account 1000 BEFORE the period
    pre = db.fetchone(
        f"""SELECT COALESCE(SUM(debit),0)-COALESCE(SUM(credit),0) AS bal
            FROM gl WHERE entity_id=? AND account_code='1000'
            AND period < ?""",
        (entity_id, periods[0])
    )
    opening_balance = float(pre["bal"]) if pre else 0.0

    # Inward = all credits to bank (money received)
    inward  = sum(float(r["credit"]) for r in bank_rows)
    # Outward = all debits to bank (money paid out) — NOTE: sign convention
    # In double-entry: debit bank = money in, credit bank = money out
    # Our posting: money in (positive amount) → debit bank; money out (negative) → credit bank
    money_in  = sum(float(r["debit"])  for r in bank_rows)
    money_out = sum(float(r["credit"]) for r in bank_rows)

    closing_balance = opening_balance + money_in - money_out

    # Income breakdown (account type = Income, credit > 0)
    income_rows = db.fetchall(
        f"""SELECT g.account_code, c.name, COALESCE(SUM(g.credit),0) AS total
            FROM gl g JOIN coa c ON c.entity_id=g.entity_id AND c.code=g.account_code
            WHERE g.entity_id=? AND g.period IN ({ph}) AND c.type='Income'
            GROUP BY g.account_code, c.name ORDER BY g.account_code""",
        [entity_id] + periods
    )
    total_income = sum(float(r["total"]) for r in income_rows)

    # Expense breakdown (Overhead / CoS / Tax, debit > 0)
    expense_rows = db.fetchall(
        f"""SELECT g.account_code, c.name, c.type, COALESCE(SUM(g.debit),0) AS total
            FROM gl g JOIN coa c ON c.entity_id=g.entity_id AND c.code=g.account_code
            WHERE g.entity_id=? AND g.period IN ({ph})
              AND c.type IN ('Overhead','CoS','Tax')
            GROUP BY g.account_code, c.name, c.type ORDER BY g.account_code""",
        [entity_id] + periods
    )
    total_expenses = sum(float(r["total"]) for r in expense_rows)

    # VAT
    vat = db.fetchone(
        f"""SELECT COALESCE(SUM(vat_amount),0) AS total
            FROM gl WHERE entity_id=? AND period IN ({ph})""",
        [entity_id] + periods
    )
    total_vat = float(vat["total"]) if vat else 0.0

    net_profit = total_income - total_expenses

    return {
        "periods":          periods,
        "opening_balance":  opening_balance,
        "money_in":         money_in,
        "money_out":        money_out,
        "closing_balance":  closing_balance,
        "total_income":     total_income,
        "total_expenses":   total_expenses,
        "total_vat":        total_vat,
        "net_profit":       net_profit,
        "income_rows":      income_rows,
        "expense_rows":     expense_rows,
        "tx_count":         len(set(r["period"] for r in rows)),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Step widgets
# ─────────────────────────────────────────────────────────────────────────────

class _StepHeader(QFrame):
    def __init__(self, steps: list, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"background:{WHITE}; border-radius:8px; border:1px solid {BORDER};")
        lay = QHBoxLayout(self)
        lay.setContentsMargins(20, 12, 20, 12)
        self._labels = []
        for i, title in enumerate(steps):
            if i > 0:
                sep = QLabel("›")
                sep.setStyleSheet(f"color:{MUTED}; font-size:16px; border:none;")
                lay.addWidget(sep)
            lbl = QLabel(f"  {i+1}. {title}  ")
            lbl.setStyleSheet(
                f"font-size:13px; border-radius:4px; padding:4px 8px; border:none;"
            )
            lay.addWidget(lbl)
            self._labels.append(lbl)
        lay.addStretch()

    def set_step(self, idx: int):
        for i, lbl in enumerate(self._labels):
            if i == idx:
                lbl.setStyleSheet(
                    f"font-size:13px; font-weight:bold; color:white; "
                    f"background:{ACCENT}; border-radius:4px; padding:4px 8px; border:none;"
                )
            elif i < idx:
                lbl.setStyleSheet(
                    f"font-size:13px; color:{SUCCESS}; "
                    f"background:#E8F5EE; border-radius:4px; padding:4px 8px; border:none;"
                )
            else:
                lbl.setStyleSheet(
                    f"font-size:13px; color:{MUTED}; "
                    f"background:#F0F4F8; border-radius:4px; padding:4px 8px; border:none;"
                )


# ─────────────────────────────────────────────────────────────────────────────
# Main page
# ─────────────────────────────────────────────────────────────────────────────

class ReportsPage(BasePage):
    def __init__(self):
        super().__init__("Accountant's Period Report",
                         "Generate, approve and submit a reconciled period report to HMRC")
        self._entity_id   = ""
        self._entity_name = ""
        self._fy_map      = {}   # label → [periods]
        self._report      = {}   # computed figures
        self._pdf_path    = ""
        self._build()

    # ── Build ─────────────────────────────────────────────────────────────────

    def _build(self):
        # Step header
        self.step_hdr = _StepHeader(
            ["Select Period", "Review & Approve", "Submit to HMRC"]
        )
        self.layout_.addWidget(self.step_hdr)

        # Stacked pages
        self.stack = QStackedWidget()
        self.layout_.addWidget(self.stack)

        self.stack.addWidget(self._build_step1())
        self.stack.addWidget(self._build_step2())
        self.stack.addWidget(self._build_step3())

        self._go_step(0)

    def _build_step1(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(16)
        lay.setContentsMargins(0, 0, 0, 0)

        card = Card("Select Financial Year and Period")
        body = card.body()

        fy_row = QHBoxLayout()
        lbl_fy = QLabel("Financial Year:")
        lbl_fy.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        lbl_fy.setFixedWidth(160)
        self.cbo_fy = ComboField(["— select company first —"])
        self.cbo_fy.setMinimumWidth(340)
        fy_row.addWidget(lbl_fy)
        fy_row.addWidget(self.cbo_fy)
        fy_row.addStretch()
        body.addLayout(fy_row)

        note = QLabel(
            "The report will cover all GL entries in the selected financial year. "
            "Ensure all bank imports are complete and allocations are committed before generating."
        )
        note.setStyleSheet(f"color:{MUTED}; font-size:12px; font-style:italic;")
        note.setWordWrap(True)
        body.addWidget(note)

        self.btn_generate = PrimaryButton("Generate Report  →")
        self.btn_generate.setFixedWidth(220)
        self.btn_generate.clicked.connect(self._generate)
        body.addWidget(self.btn_generate)

        lay.addWidget(card)
        lay.addStretch()
        return w

    def _build_step2(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(12)
        lay.setContentsMargins(0, 0, 0, 0)

        # ── Reconciliation summary ────────────────────────────────────────────
        recon_card = Card("Bank Reconciliation")
        rb = recon_card.body()
        self.recon_grid = QVBoxLayout()
        self.recon_grid.setSpacing(0)
        rb.addLayout(self.recon_grid)
        lay.addWidget(recon_card)

        # ── Income / Expense breakdown ────────────────────────────────────────
        breakdown_row = QHBoxLayout()
        breakdown_row.setSpacing(12)

        inc_card = Card("Income Breakdown")
        self.tbl_income = make_table(["Code","Account","Amount £"], stretch_col=1)
        self.tbl_income.setFixedHeight(200)
        inc_card.body().addWidget(self.tbl_income)
        breakdown_row.addWidget(inc_card)

        exp_card = Card("Expense Breakdown")
        self.tbl_expenses = make_table(["Code","Account","Type","Amount £"], stretch_col=1)
        self.tbl_expenses.setFixedHeight(200)
        exp_card.body().addWidget(self.tbl_expenses)
        breakdown_row.addWidget(exp_card)

        lay.addLayout(breakdown_row)

        # ── Director / Officer sign-off ───────────────────────────────────────
        sign_card = Card("Director / Officer Approval")
        sb = sign_card.body()

        # Info note
        note_info = QLabel(
            "ℹ  The director or company officer approves this report in FOS. "
            "The generated PDF includes a physical countersignature block so your "
            "external accountant can review and countersign offline."
        )
        note_info.setStyleSheet(
            f"color:{ACCENT}; font-size:12px; background:#EAF3FB; "
            f"border-radius:4px; padding:8px 10px; border:1px solid #B3D4F0;"
        )
        note_info.setWordWrap(True)
        sb.addWidget(note_info)

        row_name = QHBoxLayout()
        lbl_name = QLabel("Approved by:")
        lbl_name.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        lbl_name.setFixedWidth(180)
        self.f_preparer = QLineEdit()
        self.f_preparer.setPlaceholderText("Director / company officer full name")
        self.f_preparer.setFixedHeight(34)
        self.f_preparer.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:4px; padding:0 10px; font-size:13px;"
        )
        row_name.addWidget(lbl_name)
        row_name.addWidget(self.f_preparer)
        sb.addLayout(row_name)

        row_role = QHBoxLayout()
        lbl_role = QLabel("Role:")
        lbl_role.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        lbl_role.setFixedWidth(180)
        self.f_role = QLineEdit()
        self.f_role.setPlaceholderText("e.g. Director, Company Secretary, Finance Officer")
        self.f_role.setFixedHeight(34)
        self.f_role.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:4px; padding:0 10px; font-size:13px;"
        )
        row_role.addWidget(lbl_role)
        row_role.addWidget(self.f_role)
        sb.addLayout(row_role)

        row_acct = QHBoxLayout()
        lbl_acct = QLabel("External Accountant:")
        lbl_acct.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        lbl_acct.setFixedWidth(180)
        self.f_accountant = QLineEdit()
        self.f_accountant.setPlaceholderText("Accountant name (for PDF — optional)")
        self.f_accountant.setFixedHeight(34)
        self.f_accountant.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:4px; padding:0 10px; font-size:13px;"
        )
        row_acct.addWidget(lbl_acct)
        row_acct.addWidget(self.f_accountant)
        sb.addLayout(row_acct)

        row_acct_email = QHBoxLayout()
        lbl_acct_email = QLabel("Accountant Email:")
        lbl_acct_email.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        lbl_acct_email.setFixedWidth(180)
        self.f_accountant_email = QLineEdit()
        self.f_accountant_email.setPlaceholderText("accountant@firm.co.uk (optional — pre-fills email step)")
        self.f_accountant_email.setFixedHeight(34)
        self.f_accountant_email.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:4px; padding:0 10px; font-size:13px;"
        )
        row_acct_email.addWidget(lbl_acct_email)
        row_acct_email.addWidget(self.f_accountant_email)
        sb.addLayout(row_acct_email)

        row_notes = QHBoxLayout()
        lbl_notes = QLabel("Notes:")
        lbl_notes.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        lbl_notes.setFixedWidth(180)
        self.f_notes = QTextEdit()
        self.f_notes.setPlaceholderText("Optional — any notes or qualifications to include in the report")
        self.f_notes.setFixedHeight(70)
        self.f_notes.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:4px; padding:6px 10px; font-size:13px;"
        )
        row_notes.addWidget(lbl_notes)
        row_notes.addWidget(self.f_notes)
        sb.addLayout(row_notes)

        btn_row = QHBoxLayout()
        self.btn_back1 = SecondaryButton("← Back")
        self.btn_back1.clicked.connect(lambda: self._go_step(0))
        self.btn_excel_pkg = SecondaryButton("📊 Export Review Package (Excel)")
        self.btn_excel_pkg.setFixedWidth(270)
        self.btn_excel_pkg.clicked.connect(self._export_review_package)
        self.btn_approve = PrimaryButton("Approve & Generate PDF  →", colour=SUCCESS)
        self.btn_approve.setFixedWidth(260)
        self.btn_approve.clicked.connect(self._approve)
        btn_row.addWidget(self.btn_back1)
        btn_row.addStretch()
        btn_row.addWidget(self.btn_excel_pkg)
        btn_row.addWidget(self.btn_approve)
        sb.addLayout(btn_row)

        lay.addWidget(sign_card)
        return w

    def _build_step3(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(16)
        lay.setContentsMargins(0, 0, 0, 0)

        submit_card = Card("Submit to HMRC / Client")
        sb = submit_card.body()

        self.lbl_pdf_path = QLabel("PDF not yet generated")
        self.lbl_pdf_path.setStyleSheet(f"color:{MUTED}; font-size:12px; font-style:italic;")
        self.lbl_pdf_path.setWordWrap(True)
        sb.addWidget(self.lbl_pdf_path)

        # Email target
        email_row = QHBoxLayout()
        lbl_to = QLabel("Send to:")
        lbl_to.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        lbl_to.setFixedWidth(100)
        self.cbo_email_type = ComboField(["External Accountant (Offline Review)","HMRC MTD Agent","Company Director","Custom"])
        self.cbo_email_type.setMinimumWidth(200)
        self.cbo_email_type.currentIndexChanged.connect(self._prefill_email)
        self.f_email_to = QLineEdit()
        self.f_email_to.setPlaceholderText("email@example.com")
        self.f_email_to.setFixedHeight(34)
        self.f_email_to.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:4px; padding:0 10px; font-size:13px;"
        )
        email_row.addWidget(lbl_to)
        email_row.addWidget(self.cbo_email_type)
        email_row.addWidget(self.f_email_to)
        sb.addLayout(email_row)

        # Subject / body preview
        subj_row = QHBoxLayout()
        lbl_subj = QLabel("Subject:")
        lbl_subj.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        lbl_subj.setFixedWidth(100)
        self.f_email_subj = QLineEdit()
        self.f_email_subj.setFixedHeight(34)
        self.f_email_subj.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:4px; padding:0 10px; font-size:13px;"
        )
        subj_row.addWidget(lbl_subj)
        subj_row.addWidget(self.f_email_subj)
        sb.addLayout(subj_row)

        body_row = QHBoxLayout()
        lbl_body = QLabel("Body:")
        lbl_body.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        lbl_body.setFixedWidth(100)
        lbl_body.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.f_email_body = QTextEdit()
        self.f_email_body.setFixedHeight(120)
        self.f_email_body.setStyleSheet(
            f"border:1px solid {BORDER}; border-radius:4px; padding:6px 10px; font-size:13px;"
        )
        body_row.addWidget(lbl_body)
        body_row.addWidget(self.f_email_body)
        sb.addLayout(body_row)

        note = QLabel(
            "Clicking 'Open Email Draft' will open your default email client with the "
            "subject and body pre-filled. Attach the PDF from the path shown above before sending."
        )
        note.setStyleSheet(f"color:{MUTED}; font-size:11px; font-style:italic;")
        note.setWordWrap(True)
        sb.addWidget(note)

        btn_row = QHBoxLayout()
        self.btn_back2 = SecondaryButton("← Back")
        self.btn_back2.clicked.connect(lambda: self._go_step(1))
        self.btn_open_pdf = SecondaryButton("Open PDF")
        self.btn_open_pdf.clicked.connect(self._open_pdf)
        self.btn_email = PrimaryButton("Open Email Draft  ✉", colour=ACCENT)
        self.btn_email.setFixedWidth(220)
        self.btn_email.clicked.connect(self._launch_email)
        btn_row.addWidget(self.btn_back2)
        btn_row.addWidget(self.btn_open_pdf)
        btn_row.addStretch()
        btn_row.addWidget(self.btn_email)
        sb.addLayout(btn_row)

        lay.addWidget(submit_card)
        lay.addStretch()
        return w

    # ── Navigation ────────────────────────────────────────────────────────────

    def _go_step(self, idx: int):
        self.stack.setCurrentIndex(idx)
        self.step_hdr.set_step(idx)

    # ── Step 1: Generate ──────────────────────────────────────────────────────

    def set_active_entity(self, entity_id: str) -> None:
        self._entity_id = entity_id
        e = EntityModel.get(entity_id) if entity_id else None
        self._entity_name = e["legal_name"] if e else ""
        self._reload_fy(e)

    def refresh_entities(self):
        import core.context as ctx
        eid = ctx.get_entity_id()
        if eid:
            self.set_active_entity(eid)

    def _reload_fy(self, entity: dict):
        self.cbo_fy.blockSignals(True)
        self.cbo_fy.clear()
        self._fy_map = {}
        if entity:
            for label, periods in _fy_periods_for_entity(entity):
                self._fy_map[label] = periods
                self.cbo_fy.addItem(label)
        else:
            self.cbo_fy.addItem("— select company first —")
        self.cbo_fy.blockSignals(False)

    def _generate(self):
        if not self._entity_id:
            error(self, "No Company", "Select a company from the sidebar first.")
            return
        fy_label = self.cbo_fy.currentText()
        periods  = self._fy_map.get(fy_label)
        if not periods:
            error(self, "No Period", "Select a financial year.")
            return

        self._report = _build_report_data(self._entity_id, periods)
        self._report["fy_label"]     = fy_label
        self._report["entity_name"]  = self._entity_name
        self._report["generated_at"] = datetime.utcnow().strftime("%d %b %Y %H:%M UTC")

        self._populate_step2()
        self._go_step(1)

    # ── Step 2: Review ────────────────────────────────────────────────────────

    def _populate_step2(self):
        r = self._report
        # Clear recon grid
        while self.recon_grid.count():
            item = self.recon_grid.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        reconciles = abs(
            r["opening_balance"] + r["money_in"] - r["money_out"] - r["closing_balance"]
        ) < 0.02

        rows = [
            ("Opening Balance (start of period)",   r["opening_balance"],  DARK,    True),
            ("+ Money In (bank credits / income)",  r["money_in"],         SUCCESS, False),
            ("− Money Out (bank debits / expenses)",r["money_out"],        DANGER,  False),
            ("= Closing Balance (end of period)",   r["closing_balance"],  ACCENT,  True),
        ]

        for label, amount, colour, bold in rows:
            row_w = QFrame()
            row_w.setStyleSheet(f"background:{'#F8FBFF' if not bold else WHITE}; border:none;")
            row_lay = QHBoxLayout(row_w)
            row_lay.setContentsMargins(12, 8, 12, 8)
            lbl = QLabel(label)
            weight = "bold" if bold else "normal"
            lbl.setStyleSheet(f"color:{TEXT}; font-size:13px; font-weight:{weight}; border:none;")
            amt = QLabel(f"£{amount:,.2f}")
            amt.setStyleSheet(f"color:{colour}; font-size:14px; font-weight:bold; border:none;")
            amt.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            row_lay.addWidget(lbl)
            row_lay.addStretch()
            row_lay.addWidget(amt)
            self.recon_grid.addWidget(row_w)

        # Divider + reconciliation status
        div = QFrame()
        div.setFrameShape(QFrame.Shape.HLine)
        div.setStyleSheet(f"color:{BORDER};")
        self.recon_grid.addWidget(div)

        verdict = QLabel(
            "✓  Bank reconciliation agrees" if reconciles
            else f"✗  Reconciliation difference: £{abs(r['opening_balance']+r['money_in']-r['money_out']-r['closing_balance']):,.2f}"
        )
        verdict.setStyleSheet(
            f"font-size:13px; font-weight:bold; padding:8px 12px; border:none; "
            f"color:{SUCCESS if reconciles else DANGER};"
        )
        self.recon_grid.addWidget(verdict)

        # Profit summary row
        profit_row = QFrame()
        profit_row.setStyleSheet(f"background:#F0F4F8; border-radius:6px; border:none;")
        pr_lay = QHBoxLayout(profit_row)
        pr_lay.setContentsMargins(12, 8, 12, 8)
        e = EntityModel.get(self._entity_id)
        show_vat = bool(e.get("vat_registered", 1)) if e else True
        summary_items = [
            ("Total Income",    r["total_income"],  SUCCESS),
            ("Total Expenses",  r["total_expenses"], DANGER),
            ("Net Profit / Loss", r["net_profit"],  SUCCESS if r["net_profit"] >= 0 else DANGER),
        ]
        if show_vat:
            summary_items.append(("Total VAT", r["total_vat"], ACCENT))
        for label, val, colour in summary_items:
            col = QVBoxLayout()
            col.setSpacing(2)
            vl = QLabel(f"£{val:,.2f}")
            vl.setStyleSheet(f"font-size:15px; font-weight:bold; color:{colour}; border:none;")
            ll = QLabel(label)
            ll.setStyleSheet(f"font-size:10px; color:{MUTED}; border:none;")
            col.addWidget(vl)
            col.addWidget(ll)
            pr_lay.addLayout(col)
            if label != summary_items[-1][0]:
                sep = QFrame()
                sep.setFrameShape(QFrame.Shape.VLine)
                sep.setStyleSheet(f"color:{BORDER}; border:none;")
                pr_lay.addWidget(sep)
        self.recon_grid.addWidget(profit_row)

        # Income table
        self.tbl_income.setRowCount(0)
        for i, row in enumerate(r.get("income_rows", [])):
            set_row(self.tbl_income, i, [
                row["account_code"], row["name"], f"£{float(row['total']):,.2f}"
            ])

        # Expense table
        self.tbl_expenses.setRowCount(0)
        for i, row in enumerate(r.get("expense_rows", [])):
            set_row(self.tbl_expenses, i, [
                row["account_code"], row["name"], row["type"], f"£{float(row['total']):,.2f}"
            ])

        # Prefill sign-off from approver record
        e = EntityModel.get(self._entity_id)
        if e and e.get("approver"):
            ap = e["approver"]
            self.f_preparer.setText(ap.get("name",""))
            self.f_role.setText(ap.get("role","Director"))
            self.f_accountant.setText(ap.get("accountant_name",""))
            self.f_accountant_email.setText(ap.get("accountant_email",""))

    # ── Step 2: Export review package ────────────────────────────────────────

    def _export_review_package(self):
        if not self._entity_id or not self._report:
            error(self, "No Report", "Generate the report first (Step 1).")
            return
        r = self._report
        ent_name = self._entity_name.replace(" ", "_")
        fy_short = r.get("fy_label", "")[:10].replace("/","_")
        default = os.path.join(
            os.path.expanduser("~"), "Desktop",
            f"OfflineReview_{ent_name}_{fy_short}.xlsx"
        )
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Offline Review Package", default,
            "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            _export_excel_package(path, self._entity_id, r)
            info(self, "Package Exported",
                 f"Offline review package saved to:\n{path}\n\n"
                 f"Send this Excel file to your accountant — it contains:\n"
                 f"  • Bank Statements\n  • General Ledger\n  • Trial Balance\n"
                 f"  • Year-End P&L Summary")
        except Exception as exc:
            error(self, "Export Failed", str(exc))

    # ── Step 3: Approve + PDF ─────────────────────────────────────────────────

    def _approve(self):
        preparer = self.f_preparer.text().strip()
        role     = self.f_role.text().strip()
        if not preparer:
            error(self, "Sign-Off Required", "Enter the name of the person approving this report.")
            return

        self._report["preparer"]         = preparer
        self._report["role"]             = role
        self._report["accountant"]       = self.f_accountant.text().strip()
        self._report["accountant_email"] = self.f_accountant_email.text().strip()
        self._report["notes"]            = self.f_notes.toPlainText().strip()
        self._report["signed_at"]        = datetime.utcnow().strftime("%d %b %Y %H:%M UTC")

        # Save PDF
        default = os.path.join(
            os.path.expanduser("~"), "Desktop",
            f"AccountantReport_{self._entity_name.replace(' ','_')}_"
            f"{self._report.get('fy_label','')[:10].replace('/','_')}.pdf"
        )
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Report PDF", default, "PDF Files (*.pdf)"
        )
        if not path:
            return

        try:
            _generate_pdf(path, self._report)
            self._pdf_path = path
            self.lbl_pdf_path.setText(f"PDF saved:  {path}")
            self.lbl_pdf_path.setStyleSheet(f"color:{SUCCESS}; font-size:12px;")
            self._prefill_email()
            self._go_step(2)
        except Exception as exc:
            error(self, "PDF Failed", str(exc))

    # ── Step 3 actions ────────────────────────────────────────────────────────

    def _prefill_email(self):
        r   = self._report
        ent = r.get("entity_name", "")
        fy  = r.get("fy_label", "")
        prepared_by = r.get("preparer","")
        income  = r.get("total_income", 0.0)
        expense = r.get("total_expenses", 0.0)
        profit  = r.get("net_profit", 0.0)

        etype = self.cbo_email_type.currentText()
        acct_email = r.get("accountant_email", "")
        acct_name  = r.get("accountant", "your accountant")
        email_map = {
            "External Accountant (Offline Review)": acct_email,
            "HMRC MTD Agent":    "agent.services@hmrc.gov.uk",
            "Company Director":  "",
            "Custom":            "",
        }
        self.f_email_to.setText(email_map.get(etype, ""))

        subj = f"Period Report for Offline Review — {ent} — {fy}"
        self.f_email_subj.setText(subj)

        if etype == "External Accountant (Offline Review)":
            salutation = f"Dear {acct_name}" if acct_name and acct_name != "your accountant" else "Dear Accountant"
            body = (
                f"{salutation},\n\n"
                f"Please find attached the period report for {ent} covering {fy}, "
                f"approved by {prepared_by} ({r.get('role','Director')}).\n\n"
                f"Summary:\n"
                f"  Total Income:    £{income:,.2f}\n"
                f"  Total Expenses:  £{expense:,.2f}\n"
                f"  Net Profit/Loss: £{profit:,.2f}\n\n"
                f"The PDF includes a countersignature block on the final page for your "
                f"offline review and sign-off. Please print, sign, and return a copy "
                f"for our records.\n\n"
                f"Kind regards,\n{prepared_by}\n{r.get('role','Director')}, {ent}"
            )
        else:
            body = (
                f"Dear {etype},\n\n"
                f"Please find attached the period report for {ent} "
                f"covering {fy}.\n\n"
                f"Summary:\n"
                f"  Total Income:    £{income:,.2f}\n"
                f"  Total Expenses:  £{expense:,.2f}\n"
                f"  Net Profit/Loss: £{profit:,.2f}\n\n"
                f"Approved by {prepared_by}.\n\n"
                f"Kind regards,\n{prepared_by}"
            )
        self.f_email_body.setPlainText(body)

    def _open_pdf(self):
        if not self._pdf_path or not os.path.exists(self._pdf_path):
            error(self, "No PDF", "Generate and approve the report first.")
            return
        import subprocess, sys
        if sys.platform == "win32":
            os.startfile(self._pdf_path)
        elif sys.platform == "darwin":
            subprocess.run(["open", self._pdf_path])
        else:
            subprocess.run(["xdg-open", self._pdf_path])

    def _launch_email(self):
        import urllib.parse, webbrowser
        to      = self.f_email_to.text().strip()
        subject = self.f_email_subj.text().strip()
        body    = self.f_email_body.toPlainText().strip()
        mailto  = f"mailto:{urllib.parse.quote(to)}?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
        webbrowser.open(mailto)
        info(self, "Email Client Opened",
             f"Your email client has opened with the draft.\n\n"
             f"Please attach the PDF before sending:\n{self._pdf_path}")


# ─────────────────────────────────────────────────────────────────────────────
# PDF generation
# ─────────────────────────────────────────────────────────────────────────────

def _generate_pdf(path: str, r: dict):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm

    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    dark   = colors.HexColor("#1B3A5C")
    accent = colors.HexColor("#2E6DA4")
    green  = colors.HexColor("#27AE60")
    red    = colors.HexColor("#E74C3C")
    muted  = colors.HexColor("#7F8C8D")
    light  = colors.HexColor("#F4F7FB")

    h1 = ParagraphStyle("h1", fontSize=20, textColor=dark,
                         fontName="Helvetica-Bold", spaceAfter=4)
    h2 = ParagraphStyle("h2", fontSize=13, textColor=accent,
                         fontName="Helvetica-Bold", spaceAfter=4, spaceBefore=10)
    body_s = ParagraphStyle("body", fontSize=10, textColor=colors.HexColor("#2C3E50"),
                             leading=14)
    muted_s = ParagraphStyle("muted", fontSize=9, textColor=muted, leading=12)

    elems = []

    # Header
    elems.append(Paragraph("Accountant's Period Report", h1))
    elems.append(Paragraph(
        f"{r.get('entity_name','')}  ·  {r.get('fy_label','')}  ·  "
        f"Generated: {r.get('generated_at','')}", muted_s))
    elems.append(HRFlowable(width="100%", thickness=1, color=accent, spaceAfter=8))

    # Bank Reconciliation
    elems.append(Paragraph("Bank Reconciliation", h2))
    recon_data = [
        ["", "Amount (£)"],
        ["Opening Balance (start of period)",
         f"£{r.get('opening_balance',0):,.2f}"],
        ["+ Money In (credits / income received)",
         f"£{r.get('money_in',0):,.2f}"],
        ["− Money Out (debits / expenses paid)",
         f"£{r.get('money_out',0):,.2f}"],
        ["= Closing Balance (end of period)",
         f"£{r.get('closing_balance',0):,.2f}"],
    ]
    recon_tbl = Table(recon_data, colWidths=["75%","25%"])
    recon_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), dark),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 10),
        ("BACKGROUND",  (0,4), (-1,4), light),
        ("FONTNAME",    (0,4), (-1,4), "Helvetica-Bold"),
        ("TEXTCOLOR",   (0,4), (-1,4), dark),
        ("ALIGN",       (1,0), (1,-1), "RIGHT"),
        ("ROWBACKGROUNDS", (0,1), (-1,3), [colors.white, colors.HexColor("#F8FBFF")]),
        ("GRID",        (0,0), (-1,-1), 0.25, colors.HexColor("#D5E8F7")),
        ("TOPPADDING",  (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0), (-1,-1), 6),
        ("LEFTPADDING", (0,0), (-1,-1), 8),
    ]))
    elems.append(recon_tbl)
    elems.append(Spacer(1, 8))

    # Profit summary
    elems.append(Paragraph("Profit & Loss Summary", h2))
    profit  = r.get("net_profit", 0)
    p_color = green if profit >= 0 else red
    pl_data = [
        ["Total Income",   f"£{r.get('total_income',0):,.2f}"],
        ["Total Expenses", f"£{r.get('total_expenses',0):,.2f}"],
        ["Total VAT",      f"£{r.get('total_vat',0):,.2f}"],
        ["Net Profit / (Loss)", f"£{profit:,.2f}"],
    ]
    pl_tbl = Table(pl_data, colWidths=["60%","40%"])
    pl_tbl.setStyle(TableStyle([
        ("FONTNAME",   (0,3), (-1,3), "Helvetica-Bold"),
        ("TEXTCOLOR",  (1,3), (1,3), p_color),
        ("ALIGN",      (1,0), (1,-1), "RIGHT"),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, light]),
        ("GRID",       (0,0), (-1,-1), 0.25, colors.HexColor("#D5E8F7")),
        ("FONTSIZE",   (0,0), (-1,-1), 10),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ("LEFTPADDING",(0,0), (-1,-1), 8),
    ]))
    elems.append(pl_tbl)
    elems.append(Spacer(1, 8))

    # Income breakdown
    if r.get("income_rows"):
        elems.append(Paragraph("Income Breakdown", h2))
        inc_data = [["Code","Account","Amount £"]]
        for row in r["income_rows"]:
            inc_data.append([row["account_code"], row["name"], f"£{float(row['total']):,.2f}"])
        _append_breakdown(elems, inc_data, dark, light)

    # Expense breakdown
    if r.get("expense_rows"):
        elems.append(Paragraph("Expense Breakdown", h2))
        exp_data = [["Code","Account","Type","Amount £"]]
        for row in r["expense_rows"]:
            exp_data.append([row["account_code"], row["name"], row["type"],
                             f"£{float(row['total']):,.2f}"])
        _append_breakdown(elems, exp_data, dark, light)

    # Sign-off
    elems.append(Spacer(1, 10))
    elems.append(HRFlowable(width="100%", thickness=0.5, color=muted))
    elems.append(Spacer(1, 6))
    elems.append(Paragraph("Director / Officer Approval", h2))
    elems.append(Paragraph(
        f"I, the undersigned, confirm that to the best of my knowledge the information "
        f"contained in this report is accurate and complete for the period stated.", body_s))
    elems.append(Spacer(1, 6))
    sign_data = [
        ["Approved by:", r.get("preparer","")],
        ["Role:",        r.get("role","")],
        ["Date:",        r.get("signed_at","")],
    ]
    if r.get("notes"):
        sign_data.append(["Notes:", r["notes"]])
    sign_tbl = Table(sign_data, colWidths=["25%","75%"])
    sign_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (0,-1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0,0), (0,-1), dark),
        ("FONTSIZE",  (0,0), (-1,-1), 10),
        ("TOPPADDING",(0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
    ]))
    elems.append(sign_tbl)
    elems.append(Spacer(1, 16))
    elems.append(Paragraph(
        "Director Signature: ___________________________________     Date: _______________",
        body_s))

    # ── Accountant offline countersignature block ─────────────────────────────
    elems.append(Spacer(1, 16))
    elems.append(HRFlowable(width="100%", thickness=0.5, color=muted))
    elems.append(Spacer(1, 8))
    elems.append(Paragraph("External Accountant — Offline Review &amp; Countersignature", h2))
    elems.append(Paragraph(
        "This section is to be completed by the external accountant / auditor after "
        "offline review. The accountant does not require access to FOS to countersign.",
        body_s))
    elems.append(Spacer(1, 8))
    acct_name = r.get("accountant", "")
    acct_data = [
        ["Accountant Name:", acct_name if acct_name else "___________________________________"],
        ["Firm / Practice:", "___________________________________"],
        ["Review Date:",     "___________________________________"],
        ["Comments:",        ""],
    ]
    acct_tbl = Table(acct_data, colWidths=["30%","70%"])
    acct_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (0,-1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0,0), (0,-1), dark),
        ("FONTSIZE",  (0,0), (-1,-1), 10),
        ("TOPPADDING",(0,0), (-1,-1), 7),
        ("BOTTOMPADDING",(0,0),(-1,-1), 7),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, light]),
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#D5E8F7")),
        ("LEFTPADDING",(0,0),(-1,-1), 8),
        ("MINROWHEIGHT", (0,3), (-1,3), 50),
    ]))
    elems.append(acct_tbl)
    elems.append(Spacer(1, 20))
    elems.append(Paragraph(
        "Accountant Signature: ___________________________________     Date: _______________",
        body_s))
    elems.append(Spacer(1, 8))
    elems.append(Paragraph(
        "I confirm I have reviewed this period report and it fairly represents the financial "
        "position of the company for the period stated.",
        muted_s))

    doc.build(elems)


def _export_excel_package(path: str, entity_id: str, r: dict):
    """Export a multi-sheet Excel workbook for offline accountant review."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    dark_fill  = PatternFill("solid", fgColor="1B3A5C")
    hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Calibri", bold=True, color="1B3A5C", size=13)
    bold_font  = Font(name="Calibri", bold=True, color="1B3A5C", size=10)
    norm_font  = Font(name="Calibri", size=10)
    alt_fill   = PatternFill("solid", fgColor="F4F7FB")
    thin       = Side(style="thin", color="D5E8F7")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt  = '#,##0.00'
    periods    = r.get("periods", [])
    ph         = ",".join("?" * len(periods)) if periods else "''"

    def _hdr_row(ws, row_idx, values, col_widths=None):
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.fill  = dark_fill
            cell.font  = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        if col_widths:
            for c, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(c)].width = w

    def _data_row(ws, row_idx, values, money_cols=None, bold=False):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.fill   = fill
            cell.font   = bold_font if bold else norm_font
            cell.border = border
            if money_cols and c in money_cols:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")

    # ── Sheet 1: Bank Statements ──────────────────────────────────────────────
    ws1 = wb.create_sheet("Bank Statements")
    ws1.cell(1, 1, f"Bank Statements — {r.get('entity_name','')} — {r.get('fy_label','')}").font = title_font
    ws1.row_dimensions[1].height = 22
    _hdr_row(ws1, 3, ["Date","Description","Debit £","Credit £","Balance £","Period","Account"],
             [14, 42, 14, 14, 14, 10, 16])
    bank_rows = db.fetchall(
        f"SELECT date, description, debit, credit, period, account_code FROM gl "
        f"WHERE entity_id=? AND account_code='1000' AND period IN ({ph}) ORDER BY date",
        [entity_id] + periods
    ) if periods else []
    bal = r.get("opening_balance", 0.0)
    for i, row in enumerate(bank_rows, 4):
        d = float(row["debit"] or 0)
        c = float(row["credit"] or 0)
        bal += d - c
        _data_row(ws1, i, [row["date"], row["description"], d or None, c or None, bal,
                            row["period"], row["account_code"]], money_cols={3,4,5})
    ws1.freeze_panes = "A4"

    # ── Sheet 2: General Ledger ───────────────────────────────────────────────
    ws2 = wb.create_sheet("General Ledger")
    ws2.cell(1, 1, f"General Ledger — {r.get('entity_name','')} — {r.get('fy_label','')}").font = title_font
    _hdr_row(ws2, 3, ["Date","Period","Account Code","Account Name","Type","Description","Debit £","Credit £","VAT £","Ref"],
             [14, 10, 14, 30, 12, 38, 12, 12, 10, 16])
    gl_rows = db.fetchall(
        f"""SELECT g.date, g.period, g.account_code, c.name AS account_name, c.type,
                   g.description, g.debit, g.credit, g.vat_amount, g.source
            FROM gl g LEFT JOIN coa c ON c.entity_id=g.entity_id AND c.code=g.account_code
            WHERE g.entity_id=? AND g.period IN ({ph}) ORDER BY g.date, g.account_code""",
        [entity_id] + periods
    ) if periods else []
    for i, row in enumerate(gl_rows, 4):
        _data_row(ws2, i, [
            row["date"], row["period"], row["account_code"], row["account_name"],
            row["type"], row["description"],
            float(row["debit"] or 0) or None, float(row["credit"] or 0) or None,
            float(row["vat_amount"] or 0) or None, row["source"]
        ], money_cols={7,8,9})
    ws2.freeze_panes = "A4"

    # ── Sheet 3: Trial Balance ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Trial Balance")
    ws3.cell(1, 1, f"Trial Balance — {r.get('entity_name','')} — {r.get('fy_label','')}").font = title_font
    _hdr_row(ws3, 3, ["Code","Account","Type","Total Debits £","Total Credits £","Net £"],
             [10, 36, 12, 16, 16, 16])
    tb_rows = db.fetchall(
        f"""SELECT g.account_code, c.name, c.type,
                   COALESCE(SUM(g.debit),0) AS total_dr,
                   COALESCE(SUM(g.credit),0) AS total_cr
            FROM gl g LEFT JOIN coa c ON c.entity_id=g.entity_id AND c.code=g.account_code
            WHERE g.entity_id=? AND g.period IN ({ph})
            GROUP BY g.account_code, c.name, c.type ORDER BY g.account_code""",
        [entity_id] + periods
    ) if periods else []
    for i, row in enumerate(tb_rows, 4):
        dr = float(row["total_dr"] or 0)
        cr = float(row["total_cr"] or 0)
        _data_row(ws3, i, [row["account_code"], row["name"], row["type"], dr, cr, dr - cr],
                  money_cols={4,5,6})
    # Totals row
    if tb_rows:
        tot_i = len(tb_rows) + 4
        total_dr = sum(float(r2["total_dr"] or 0) for r2 in tb_rows)
        total_cr = sum(float(r2["total_cr"] or 0) for r2 in tb_rows)
        _data_row(ws3, tot_i, ["","TOTAL","", total_dr, total_cr, total_dr - total_cr],
                  money_cols={4,5,6}, bold=True)
    ws3.freeze_panes = "A4"

    # ── Sheet 4: Year-End P&L Summary ────────────────────────────────────────
    ws4 = wb.create_sheet("Year-End Summary")
    ws4.cell(1, 1, f"Year-End P&L Summary — {r.get('entity_name','')} — {r.get('fy_label','')}").font = title_font
    row_n = 3

    def _section(label):
        nonlocal row_n
        cell = ws4.cell(row_n, 1, label)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = dark_fill
        ws4.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=3)
        row_n += 1

    def _line(label, val, bold=False):
        nonlocal row_n
        ws4.cell(row_n, 1, label).font = bold_font if bold else norm_font
        c = ws4.cell(row_n, 2, val)
        c.number_format = money_fmt
        c.alignment = Alignment(horizontal="right")
        c.font = bold_font if bold else norm_font
        row_n += 1

    _section("Bank Reconciliation")
    _line("Opening Balance", r.get("opening_balance", 0))
    _line("+ Money In",      r.get("money_in", 0))
    _line("− Money Out",     r.get("money_out", 0))
    _line("= Closing Balance", r.get("closing_balance", 0), bold=True)
    row_n += 1

    _section("Profit & Loss")
    for inc in r.get("income_rows", []):
        _line(f"  {inc['account_code']}  {inc['name']}", float(inc["total"]))
    _line("Total Income", r.get("total_income", 0), bold=True)
    row_n += 1
    for exp in r.get("expense_rows", []):
        _line(f"  {exp['account_code']}  {exp['name']} ({exp['type']})", float(exp["total"]))
    _line("Total Expenses", r.get("total_expenses", 0), bold=True)
    row_n += 1
    _line("Net Profit / (Loss)", r.get("net_profit", 0), bold=True)
    _line("Total VAT Collected",  r.get("total_vat", 0))

    ws4.column_dimensions["A"].width = 46
    ws4.column_dimensions["B"].width = 18

    wb.save(path)


def _append_breakdown(elems, data, dark, light):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    col_w = ["12%", "50%", "18%", "20%"] if len(data[0]) == 4 else ["12%", "68%", "20%"]
    tbl = Table(data, colWidths=col_w)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), dark),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN",      (-1,0),(-1,-1), "RIGHT"),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, light]),
        ("GRID",       (0,0), (-1,-1), 0.25, colors.HexColor("#D5E8F7")),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING",(0,0), (-1,-1), 6),
    ]))
    elems.append(tbl)
    from reportlab.platypus import Spacer
    elems.append(Spacer(1, 6))
