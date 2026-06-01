"""
fos/ui/page_gl.py
General Ledger and Trial Balance viewer.
"""

from PyQt6.QtWidgets import QHBoxLayout, QVBoxLayout, QLabel, QWidget
from PyQt6.QtCore import Qt

from ui.widgets import (
    BasePage, Card, SecondaryButton, ComboField,
    make_table, set_row,
    ACCENT, DARK, TEXT, MUTED, WHITE, BG, SUCCESS, DANGER, BORDER
)
from core.models import EntityModel, GLModel
from core.database import db

_MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun",
               "Jul","Aug","Sep","Oct","Nov","Dec"]
_MONTH_NAMES = ["January","February","March","April","May","June",
                "July","August","September","October","November","December"]


def _fy_options(entity: dict) -> list:
    """Return list of (label, start_YYYY-MM, end_YYYY-MM) for recent FY years."""
    try:
        fy_start_name = entity.get("fy_start", "April")
        fy_start_mo   = (_MONTH_NAMES.index(fy_start_name) + 1) if fy_start_name in _MONTH_NAMES else 4
    except Exception:
        fy_start_mo = 4

    from datetime import date
    today = date.today()
    options = []
    for delta in range(3):  # current + 2 prior FYs
        # FY that started `delta` years ago
        start_year = today.year - delta
        if today.month < fy_start_mo:
            start_year -= 1
        end_year = start_year + 1
        start_str = f"{start_year}-{fy_start_mo:02d}"
        # end period = month before fy_start in end_year
        end_mo = fy_start_mo - 1 or 12
        end_yr = end_year if fy_start_mo > 1 else start_year
        end_str = f"{end_yr}-{end_mo:02d}"
        label = (f"FY {start_year}/{str(end_year)[2:]}  "
                 f"({_MONTH_ABBR[fy_start_mo-1]} {start_year} – "
                 f"{_MONTH_ABBR[end_mo-1]} {end_yr})")
        options.append((label, start_str, end_str))
    return options


def _periods_in_range(start: str, end: str) -> list:
    """Return list of YYYY-MM strings between start and end inclusive."""
    from datetime import date
    sy, sm = int(start[:4]), int(start[5:7])
    ey, em = int(end[:4]),   int(end[5:7])
    result = []
    y, m = sy, sm
    while (y, m) <= (ey, em):
        result.append(f"{y}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return result


def _get_entries_filtered(entity_id: str, fy_periods, month_period) -> list:
    if month_period:
        return GLModel.get_entries(entity_id, month_period)
    if fy_periods:
        periods = _periods_in_range(*fy_periods)
        placeholders = ",".join("?" * len(periods))
        return db.fetchall(
            f"SELECT * FROM gl WHERE entity_id=? AND period IN ({placeholders}) "
            f"ORDER BY date, gl_id",
            [entity_id] + periods
        )
    return GLModel.get_entries(entity_id, None)


def _get_tb_filtered(entity_id: str, fy_periods) -> list:
    if fy_periods:
        periods = _periods_in_range(*fy_periods)
        placeholders = ",".join("?" * len(periods))
        return db.fetchall(
            f"""SELECT c.code, c.name, c.type, c.normal_balance,
                       COALESCE(SUM(g.debit),0)  AS total_debit,
                       COALESCE(SUM(g.credit),0) AS total_credit,
                       COALESCE(SUM(g.debit),0) - COALESCE(SUM(g.credit),0) AS balance
                FROM coa c
                LEFT JOIN gl g ON g.entity_id=c.entity_id AND g.account_code=c.code
                    AND g.period IN ({placeholders})
                WHERE c.entity_id=?
                GROUP BY c.code, c.name, c.type, c.normal_balance
                HAVING total_debit != 0 OR total_credit != 0
                ORDER BY c.code""",
            periods + [entity_id]
        )
    return GLModel.get_trial_balance(entity_id, None)


def _export_table(parent, rows: list, headers: list, fmt: str, default_name: str):
    from PyQt6.QtWidgets import QFileDialog, QMessageBox
    ext_map = {"CSV": "csv", "Excel": "xlsx", "PDF": "pdf"}
    ext = ext_map.get(fmt, "csv")
    path, _ = QFileDialog.getSaveFileName(
        parent, f"Export {fmt}", default_name, f"{fmt} Files (*.{ext})"
    )
    if not path:
        return
    try:
        if fmt == "CSV":
            _export_csv(path, rows, headers)
        elif fmt == "Excel":
            _export_excel(path, rows, headers, default_name)
        elif fmt == "PDF":
            _export_pdf(path, rows, headers, default_name)
        QMessageBox.information(parent, "Exported", f"Saved to:\n{path}")
    except Exception as exc:
        QMessageBox.critical(parent, "Export Failed", str(exc))


def _export_csv(path: str, rows: list, headers: list):
    import csv
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow(list(r.values()) if isinstance(r, dict) else list(r))


def _export_excel(path: str, rows: list, headers: list, title: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    hdr_fill = PatternFill("solid", fgColor="1B3A5C")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for row_idx, r in enumerate(rows, 2):
        vals = list(r.values()) if isinstance(r, dict) else list(r)
        for col, v in enumerate(vals, 1):
            ws.cell(row=row_idx, column=col, value=v)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(col[0].value or "")), 12
        ) + 2
    wb.save(path)


def _export_pdf(path: str, rows: list, headers: list, title: str):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles["Title"]),
        Spacer(1, 12),
    ]
    data = [headers]
    for r in rows:
        vals = list(r.values()) if isinstance(r, dict) else list(r)
        data.append([str(v or "") for v in vals])
    col_count = len(headers)
    avail_width = landscape(A4)[0] - 40
    col_w = avail_width / col_count
    tbl = Table(data, colWidths=[col_w] * col_count, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1B3A5C")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F4F7FB")]),
        ("GRID",       (0,0), (-1,-1), 0.25, colors.HexColor("#D5E8F7")),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
    ]))
    elements.append(tbl)
    doc.build(elements)


class GLPage(BasePage):
    def __init__(self):
        super().__init__("General Ledger",
                         "View all committed transactions by account")
        self._entity_map: dict = {}
        self._fy_map: dict = {}
        self._gl_entries: list = []
        self._build()

    def _build(self):
        bar = QHBoxLayout()
        lbl = QLabel("Company:")
        lbl.setVisible(False)
        self.cbo_entity = ComboField(["— select company —"])
        self.cbo_entity.setMinimumWidth(260)
        self.cbo_entity.setVisible(False)
        self.cbo_entity.currentIndexChanged.connect(self._on_entity_change)

        lbl_fy = QLabel("FY:")
        lbl_fy.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        self.cbo_fy = ComboField(["All Years"])
        self.cbo_fy.setMinimumWidth(180)
        self.cbo_fy.currentIndexChanged.connect(self._load)

        lbl_mo = QLabel("  Month:")
        lbl_mo.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        self.cbo_period = ComboField(["All Periods"])
        self.cbo_period.setMinimumWidth(120)
        self.cbo_period.currentIndexChanged.connect(self._load)

        self.btn_refresh = SecondaryButton("⟳  Refresh")
        self.btn_refresh.clicked.connect(self._load)
        bar.addWidget(lbl)
        bar.addWidget(self.cbo_entity)
        bar.addWidget(lbl_fy)
        bar.addWidget(self.cbo_fy)
        bar.addWidget(lbl_mo)
        bar.addWidget(self.cbo_period)
        bar.addStretch()
        bar.addWidget(self.btn_refresh)
        self.layout_.addLayout(bar)

        # Summary
        self.lbl_summary = QLabel("")
        self.lbl_summary.setStyleSheet(f"color:{MUTED}; font-size:12px;")
        self.layout_.addWidget(self.lbl_summary)

        gl_card = Card("GL Entries")
        self.tbl = make_table(
            ["Date","Account","Description","Debit £","Credit £","VAT","Source","Period"],
            stretch_col=2
        )
        self.tbl.setMinimumHeight(450)
        self.tbl.setColumnWidth(0, 90)
        self.tbl.setColumnWidth(1, 120)
        self.tbl.setColumnWidth(3, 90)
        self.tbl.setColumnWidth(4, 90)
        self.tbl.setColumnWidth(5, 60)
        self.tbl.setColumnWidth(6, 90)
        self.tbl.setColumnWidth(7, 70)
        gl_card.body().addWidget(self.tbl)
        self.layout_.addWidget(gl_card)

        exp_row = QHBoxLayout()
        exp_row.addStretch()
        for fmt in ["CSV", "Excel", "PDF"]:
            btn = SecondaryButton(f"⬇  Export {fmt}")
            btn.clicked.connect(lambda _, f=fmt: self._export(f))
            exp_row.addWidget(btn)
        self.layout_.addLayout(exp_row)
        self.layout_.addStretch()
        self.refresh_entities()

    def refresh_entities(self):
        self.cbo_entity.blockSignals(True)
        self.cbo_entity.clear()
        self._entity_map = {}
        for e in EntityModel.list_all():
            self.cbo_entity.addItem(e["legal_name"])
            self._entity_map[e["legal_name"]] = e["entity_id"]
        if not self._entity_map:
            self.cbo_entity.addItem("— no companies yet —")
        self.cbo_entity.blockSignals(False)
        self._on_entity_change()

    def set_active_entity(self, entity_id: str) -> None:
        for name, eid in self._entity_map.items():
            if eid == entity_id:
                self.cbo_entity.blockSignals(True)
                self.cbo_entity.setCurrentText(name)
                self.cbo_entity.blockSignals(False)
                break
        self._on_entity_change()

    def _on_entity_change(self):
        import core.context as ctx
        entity_id = ctx.get_entity_id() or self._entity_map.get(self.cbo_entity.currentText(), "")
        self._refresh_fy(entity_id)
        self._load()

    def _refresh_fy(self, entity_id: str):
        self._fy_map = {}
        self.cbo_fy.blockSignals(True)
        self.cbo_fy.clear()
        self.cbo_fy.addItem("All Years")
        if entity_id:
            e = EntityModel.get(entity_id)
            if e:
                for label, start, end in _fy_options(e):
                    self._fy_map[label] = (start, end)
                    self.cbo_fy.addItem(label)
        self.cbo_fy.blockSignals(False)

    def _load(self):
        import core.context as ctx
        entity_id = ctx.get_entity_id() or self._entity_map.get(self.cbo_entity.currentText(),"")
        if not entity_id:
            return

        # Determine period filter from FY + month selections
        fy_label = self.cbo_fy.currentText()
        fy_periods = None
        if fy_label != "All Years" and fy_label in self._fy_map:
            fy_periods = self._fy_map[fy_label]  # (start_YYYY-MM, end_YYYY-MM)

        period = self.cbo_period.currentText()
        if period == "All Periods":
            period = None

        entries = _get_entries_filtered(entity_id, fy_periods, period)

        # Refresh month dropdown
        all_periods = db.fetchall(
            "SELECT DISTINCT period FROM gl WHERE entity_id=? ORDER BY period DESC",
            (entity_id,)
        )
        self.cbo_period.blockSignals(True)
        cur = self.cbo_period.currentText()
        self.cbo_period.clear()
        self.cbo_period.addItem("All Periods")
        for p in all_periods:
            self.cbo_period.addItem(p["period"])
        idx = self.cbo_period.findText(cur)
        if idx >= 0:
            self.cbo_period.setCurrentIndex(idx)
        self.cbo_period.blockSignals(False)

        self.tbl.setRowCount(0)
        total_debit = 0.0
        total_credit = 0.0
        for i, e in enumerate(entries):
            debit  = float(e.get("debit",0))
            credit = float(e.get("credit",0))
            total_debit  += debit
            total_credit += credit
            set_row(self.tbl, i, [
                e["date"][:10],
                e["account_code"],
                (e.get("description") or "")[:70],
                f"{debit:,.2f}"  if debit  else "",
                f"{credit:,.2f}" if credit else "",
                e.get("vat_code",""),
                e.get("source",""),
                e.get("period",""),
            ])

        diff = total_debit - total_credit
        colour = SUCCESS if abs(diff) < 0.01 else DANGER
        self.lbl_summary.setText(
            f"Entries: {len(entries)}    "
            f"Total Debit: £{total_debit:,.2f}    "
            f"Total Credit: £{total_credit:,.2f}    "
            f"Difference: £{diff:,.2f}"
        )
        self.lbl_summary.setStyleSheet(f"color:{colour}; font-size:12px; font-weight:bold;")
        self._gl_entries = entries

    def _export(self, fmt: str):
        company = self.cbo_entity.currentText()
        fy = self.cbo_fy.currentText().replace("/", "-")
        default_name = f"GeneralLedger_{company}_{fy}"
        _export_table(self, self._gl_entries,
                      ["Date","Account","Description","Debit £","Credit £","VAT","Source","Period"],
                      fmt, default_name)


class TrialBalancePage(BasePage):
    def __init__(self):
        super().__init__("Trial Balance",
                         "Debit and credit balances by account — must agree before year-end")
        self._entity_map: dict = {}
        self._fy_map: dict = {}
        self._tb_rows: list = []
        self._build()

    def _build(self):
        bar = QHBoxLayout()
        lbl = QLabel("Company:")
        lbl.setVisible(False)
        self.cbo_entity = ComboField(["— select company —"])
        self.cbo_entity.setMinimumWidth(260)
        self.cbo_entity.setVisible(False)
        self.cbo_entity.currentIndexChanged.connect(self._on_entity_change)

        lbl_fy = QLabel("FY:")
        lbl_fy.setStyleSheet(f"color:{TEXT}; font-size:13px;")
        self.cbo_fy = ComboField(["All Years"])
        self.cbo_fy.setMinimumWidth(180)
        self.cbo_fy.currentIndexChanged.connect(self._load)

        self.btn_refresh = SecondaryButton("⟳  Refresh")
        self.btn_refresh.clicked.connect(self._load)
        bar.addWidget(lbl)
        bar.addWidget(self.cbo_entity)
        bar.addWidget(lbl_fy)
        bar.addWidget(self.cbo_fy)
        bar.addStretch()
        bar.addWidget(self.btn_refresh)
        self.layout_.addLayout(bar)

        self.lbl_agree = QLabel("")
        self.lbl_agree.setStyleSheet("font-size:14px; font-weight:bold;")
        self.layout_.addWidget(self.lbl_agree)

        tb_card = Card("Trial Balance")
        self.tbl = make_table(
            ["Code","Account","Type","Normal Bal.","Total Debit £","Total Credit £","Balance £"],
            stretch_col=1
        )
        self.tbl.setMinimumHeight(450)
        self.tbl.setColumnWidth(0, 70)
        self.tbl.setColumnWidth(2, 100)
        self.tbl.setColumnWidth(3, 90)
        self.tbl.setColumnWidth(4, 110)
        self.tbl.setColumnWidth(5, 110)
        self.tbl.setColumnWidth(6, 110)
        tb_card.body().addWidget(self.tbl)
        self.layout_.addWidget(tb_card)

        exp_row = QHBoxLayout()
        exp_row.addStretch()
        for fmt in ["CSV", "Excel", "PDF"]:
            btn = SecondaryButton(f"⬇  Export {fmt}")
            btn.clicked.connect(lambda _, f=fmt: self._export(f))
            exp_row.addWidget(btn)
        self.layout_.addLayout(exp_row)
        self.layout_.addStretch()
        self.refresh_entities()

    def refresh_entities(self):
        self.cbo_entity.blockSignals(True)
        self.cbo_entity.clear()
        self._entity_map = {}
        for e in EntityModel.list_all():
            self.cbo_entity.addItem(e["legal_name"])
            self._entity_map[e["legal_name"]] = e["entity_id"]
        if not self._entity_map:
            self.cbo_entity.addItem("— no companies yet —")
        self.cbo_entity.blockSignals(False)
        self._on_entity_change()

    def set_active_entity(self, entity_id: str) -> None:
        for name, eid in self._entity_map.items():
            if eid == entity_id:
                self.cbo_entity.blockSignals(True)
                self.cbo_entity.setCurrentText(name)
                self.cbo_entity.blockSignals(False)
                break
        self._on_entity_change()

    def _on_entity_change(self):
        import core.context as ctx
        entity_id = ctx.get_entity_id() or self._entity_map.get(self.cbo_entity.currentText(), "")
        self._fy_map = {}
        self.cbo_fy.blockSignals(True)
        self.cbo_fy.clear()
        self.cbo_fy.addItem("All Years")
        if entity_id:
            e = EntityModel.get(entity_id)
            if e:
                for label, start, end in _fy_options(e):
                    self._fy_map[label] = (start, end)
                    self.cbo_fy.addItem(label)
        self.cbo_fy.blockSignals(False)
        self._load()

    def _load(self):
        import core.context as ctx
        entity_id = ctx.get_entity_id() or self._entity_map.get(self.cbo_entity.currentText(),"")
        if not entity_id:
            return

        fy_label = self.cbo_fy.currentText()
        fy_periods = None
        if fy_label != "All Years" and fy_label in self._fy_map:
            fy_periods = self._fy_map[fy_label]

        self._tb_rows = _get_tb_filtered(entity_id, fy_periods)

        self.tbl.setRowCount(0)
        total_dr = 0.0
        total_cr = 0.0
        for i, r in enumerate(self._tb_rows):
            dr   = float(r.get("total_debit",0))
            cr   = float(r.get("total_credit",0))
            bal  = float(r.get("balance",0))
            total_dr += dr
            total_cr += cr
            set_row(self.tbl, i, [
                r["code"], r["name"], r["type"], r["normal_balance"],
                f"{dr:,.2f}", f"{cr:,.2f}",
                f"{bal:,.2f}",
            ])

        diff = total_dr - total_cr
        agrees = abs(diff) < 0.01
        if agrees:
            self.lbl_agree.setText(
                f"✓  Trial Balance agrees — Total Debit = Total Credit = £{total_dr:,.2f}"
            )
            self.lbl_agree.setStyleSheet(f"font-size:14px; font-weight:bold; color:{SUCCESS};")
        else:
            self.lbl_agree.setText(
                f"✗  Trial Balance does NOT agree — Difference: £{diff:,.2f}"
            )
            self.lbl_agree.setStyleSheet(f"font-size:14px; font-weight:bold; color:{DANGER};")

    def _export(self, fmt: str):
        from PyQt6.QtWidgets import QFileDialog
        company = self.cbo_entity.currentText()
        fy = self.cbo_fy.currentText().replace("/", "-")
        default_name = f"TrialBalance_{company}_{fy}"
        _export_table(self, self._tb_rows,
                      ["Code","Account","Type","Normal Bal","Debit £","Credit £","Balance £"],
                      fmt, default_name)
